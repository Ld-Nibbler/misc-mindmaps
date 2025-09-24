#!/usr/bin/env python3
"""
xlsx_to_html_table.py

Simple script: read an .xlsx and produce pure HTML <table>...</table> code.

Usage:
    python xlsx_to_html_table.py input.xlsx            # prints table to stdout (uses first sheet)
    python xlsx_to_html_table.py input.xlsx -s Sheet1  # use sheet named "Sheet1"
    python xlsx_to_html_table.py input.xlsx -n         # treat file as no-header (all rows in <tbody>)
    python xlsx_to_html_table.py input.xlsx -o out.html
"""

import argparse
import sys
from openpyxl import load_workbook
from datetime import datetime, date
import html

def cell_to_text(cell_value):
    """Convert common excel cell values to a safe string for HTML."""
    if cell_value is None:
        return ""
    if isinstance(cell_value, (datetime, date)):
        # ISO-like representation (you can change format if you prefer)
        return cell_value.isoformat()
    # for floats/ints/bools just str()
    return str(cell_value)

def sheet_to_html_table(ws, header=True):
    """
    Convert an openpyxl worksheet to a pure HTML table string.
    - header=True: first non-empty row will be used as <thead>.
    """
    rows = list(ws.iter_rows(values_only=True))
    # If sheet is empty
    if not rows:
        return "<table></table>"

    # find first non-empty row index (helpful if there are leading blank rows)
    first_non_empty = 0
    for i, r in enumerate(rows):
        if any(cell is not None and str(cell).strip() != "" for cell in r):
            first_non_empty = i
            break

    html_parts = []
    html_parts.append("<table>")

    start_body = first_non_empty
    if header:
        header_row = rows[first_non_empty]
        html_parts.append("  <thead>")
        html_parts.append("    <tr>")
        for cell in header_row:
            text = html.escape(cell_to_text(cell))
            html_parts.append(f"      <th>{text}</th>")
        html_parts.append("    </tr>")
        html_parts.append("  </thead>")
        start_body = first_non_empty + 1

    html_parts.append("  <tbody>")
    for r in rows[start_body:]:
        # skip completely empty trailing rows
        if all(cell is None or str(cell).strip() == "" for cell in r):
            continue
        html_parts.append("    <tr>")
        for cell in r:
            text = html.escape(cell_to_text(cell))
            html_parts.append(f"      <td>{text}</td>")
        html_parts.append("    </tr>")
    html_parts.append("  </tbody>")

    html_parts.append("</table>")
    return "\n".join(html_parts)

def main(argv):
    p = argparse.ArgumentParser(description="Convert .xlsx worksheet to pure HTML <table> markup.")
    p.add_argument("xlsx", help="Input .xlsx file")
    p.add_argument("-s", "--sheet", help="Sheet name (default: first sheet)", default=None)
    p.add_argument("-n", "--no-header", help="Treat the sheet as having no header row", action="store_true")
    p.add_argument("-o", "--output", help="Output file (default: stdout)", default=None)
    args = p.parse_args(argv)

    # load workbook with cached values (data_only=True), so we get evaluated values if present
    wb = load_workbook(filename=args.xlsx, data_only=True)
    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print(f"Error: sheet '{args.sheet}' not found. Available sheets: {', '.join(wb.sheetnames)}", file=sys.stderr)
            sys.exit(2)
        ws = wb[args.sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    html_table = sheet_to_html_table(ws, header=not args.no_header)

    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(html_table)
        print(f"Wrote HTML table to {args.output}")
    else:
        print(html_table)

if __name__ == "__main__":
    main(sys.argv[1:])
